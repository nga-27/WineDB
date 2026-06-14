""" Handler function for syncing spreadsheet with cloud location """
import os
from datetime import datetime
from typing import List
import time
import logging
import json
import shutil

import pandas as pd
import requests
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from terminal_ui_lite import TerminalUILite

from app.logging_config import LOGGER_NAME
from app.db.database import (
    WineSupply,
    WineType,
    Region,
    Country,
    PhysicalLocation,
    GrapeVariety,
    FoodPairing,
    Keywords,
)
from app.spreadsheet.generator import TAB_MAP
from app.db.database import SETTINGS_FILE_PATH

# pylint: disable=too-many-locals,too-many-branches,too-many-statements,broad-exception-caught


def sync_handler(ui_manager: TerminalUILite) -> bool:
    """sync_handler

    Syncs the xlsx db file with the cloud location

    Args:
        ui_manager (TerminalUILite): ui manager instance

    Returns:
        bool: on success of syncing with cloud location
    """
    logger = logging.getLogger(LOGGER_NAME)
    ui_manager.add_text_content("Starting spreadsheet sync...")

    # Get all wine supplies from database
    response = requests.get("http://localhost:8282/wine_supplies/joined", timeout=10)
    if response.status_code != 200:
        logger.error(
            "Failed to retrieve wine supplies from API. Status code: %s, Response: %s",
            response.status_code, response.text)
        ui_manager.add_text_content(
            "\033[31mError retrieving wine supplies from database. See logs for details.\033[39m")
        time.sleep(2)
        return False
    wines_data: List[dict] = response.json()
    logger.info("Retrieved %d wine records from API.", len(wines_data))
    logger.info("Wines: %s", wines_data)
    response = requests.get("http://localhost:8282/grape_varieties", timeout=10)
    if response.status_code != 200:
        logger.error(
            "Failed to retrieve grape varieties from API. Status code: %s, Response: %s",
            response.status_code, response.text)
        ui_manager.add_text_content(
            "\033[31mError retrieving grape varieties from database. See logs for details.\033[39m")
        time.sleep(2)
        return False
    grapes_data: List[dict] = response.json()
    logger.info("Retrieved %d grape variety records from API.", len(grapes_data))
    logger.info("Grapes: %s", grapes_data)
    grapes = [GrapeVariety.model_validate(g) for g in grapes_data]
    response = requests.get("http://localhost:8282/regions", timeout=10)
    if response.status_code != 200:
        logger.error(
            "Failed to retrieve regions from API. Status code: %s, Response: %s",
            response.status_code, response.text)
        ui_manager.add_text_content(
            "\033[31mError retrieving regions from database. See logs for details.\033[39m")
        time.sleep(2)
        return False
    regions_data: List[dict] = response.json()
    logger.info("Retrieved %d region records from API.", len(regions_data))
    logger.info("Regions: %s", regions_data)
    regions = [Region.model_validate(r) for r in regions_data]
    response = requests.get("http://localhost:8282/locations", timeout=10)
    if response.status_code != 200:
        logger.error(
            "Failed to retrieve locations from API. Status code: %s, Response: %s",
            response.status_code, response.text)
        ui_manager.add_text_content(
            "\033[31mError retrieving locations from database. See logs for details.\033[39m")
        time.sleep(2)
        return False
    locations_data: List[dict] = response.json()
    logger.info("Retrieved %d location records from API.", len(locations_data))
    logger.info("Locations: %s", locations_data)
    locations = [PhysicalLocation.model_validate(l) for l in locations_data]

    wines: List[WineSupply] = []
    for wine_data in wines_data:
        wine = WineSupply.model_validate(wine_data)

        # Reconstruct relationship objects from nested data
        if wine_data.get("wine_type"):
            wine.wine_type = WineType(name=wine_data["wine_type"])
        if wine_data.get("region"):
            wine.region = Region(name=wine_data["region"])
        if wine_data.get("country"):
            wine.country = Country(name=wine_data["country"])
        if wine_data.get("physical_location"):
            wine.physical_location = PhysicalLocation(name=wine_data["physical_location"])
        if wine_data.get("grapes"):
            wine.grapes = [GrapeVariety(name=g) for g in wine_data["grapes"]]
        if wine_data.get("food_pairings"):
            wine.food_pairings = [FoodPairing(name=fp) for fp in wine_data["food_pairings"]]
        if wine_data.get("keywords"):
            wine.keywords = [Keywords(keyword=k) for k in wine_data["keywords"]]
        wines.append(wine)

    logger.info("Wines after conversion: %s", wines)
    if not wines or len(wines) == 0:
        ui_manager.add_text_content("\033[31mNo wines found in database to export.\033[39m")
        time.sleep(2)
        return True

    ui_manager.add_text_content(f"Found {len(wines)} wines to export.")

    # Generate output filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), "output")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, f"WineDB_Export_{timestamp}.xlsx")

    # Create Excel writer and generate each tab
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for tab_name, tab_generator in TAB_MAP.items():
            ui_manager.add_text_content(f"Generating '{tab_name}' tab...")
            if tab_name == "Grape Varieties":
                tab_data = tab_generator(wines, grapes)
            elif tab_name == "Regions":
                tab_data = tab_generator(wines, regions)
            elif tab_name == "Locations":
                tab_data = tab_generator(wines, locations)
            else:
                tab_data = tab_generator(wines)

            if tab_data:
                logger.info("Tab data for '%s': %s", tab_name, tab_data)
                try:
                    df = pd.DataFrame(tab_data)
                except Exception as exc:
                    logger.error("Error creating DataFrame for tab '%s': %s", tab_name, str(exc))
                    ui_manager.add_text_content(
                        f"\033[31mError generating '{tab_name}' tab. See logs for details.\033[39m")
                    time.sleep(2)
                    continue
            else:
                ui_manager.add_text_content(f"\033[33mTab '{tab_name}' returned no data.\033[39m")
                continue

            # Sanitize sheet name (Excel limit: 31 chars, no special chars)
            safe_sheet_name = tab_name[:31].replace("/", "-").replace("\\", "-")
            df.to_excel(writer, sheet_name=safe_sheet_name, index=False)

            # Apply bold formatting to header-like rows
            workbook = writer.book
            sheet = workbook[safe_sheet_name]

            # Get column names
            columns = list(df.columns)

            for row_idx in range(2, sheet.max_row + 1):  # Start from row 2 (after header)
                name_cell = sheet.cell(row=row_idx, column=1)  # Assuming "Name" is first column
                if name_cell.value and all(
                    sheet.cell(row=row_idx, column=col_idx + 1).value in [None, '']
                    for col_idx in range(2, len(columns))
                ):
                    # This is a header row (Name has value, others are empty)
                    name_cell.font = Font(bold=True)
                    second_cell = sheet.cell(row=row_idx, column=2)
                    if second_cell.value:
                        second_cell.font = Font(italic=True)

            # Adjust column widths based on max content length
            for col_idx, column in enumerate(columns, start=1):
                max_length = len(str(column))  # Start with column header length
                for row_idx in range(2, sheet.max_row + 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    if cell_value is None:
                        continue
                    # Replace tabs with spaces and trim
                    cell_str = str(cell_value).replace('\t', '    ').strip()
                    # Consider multi-line cells: use the longest line length
                    lines = cell_str.splitlines() or [cell_str]
                    longest_line = max((len(line.strip()) for line in lines), default=0)
                    max_length = max(max_length, longest_line)
                # Set column width with 2-char padding for readability and clamp to a max
                width = max_length + 2
                col_letter = get_column_letter(col_idx)
                sheet.column_dimensions[col_letter].width = width

            ui_manager.add_text_content(
                f"\033[32m'{tab_name}' tab generated with {len(tab_data)} records.\033[39m")

    ui_manager.add_text_content(f"Spreadsheet exported to: {output_path}")
    if not os.path.exists(SETTINGS_FILE_PATH):
        logger.warning("Settings file not found at %s. Skipping drive sync.", SETTINGS_FILE_PATH)
        ui_manager.add_text_content("\033[33mSettings file not found. Skipping drive sync.\033[39m")
    else:
        settings = {}
        try:
            with open(SETTINGS_FILE_PATH, "r", encoding="utf-8") as file_x:
                settings = json.load(file_x)
            sync_path = settings.get("sync_path")
            if not sync_path or len(sync_path) == 0:
                logger.warning(
                    "Sync path not found or not filled in in settings. Skipping drive sync.")
                ui_manager.add_text_content(
                    "\033[33mSync path not found in settings. Skipping drive sync.\033[39m")
            else:
                shutil.copy(output_path, sync_path)
                db_output_path = os.path.join(os.path.dirname(
                    os.path.dirname(output_path)), "wineDB.db")
                db_sync_path = os.path.join(os.path.dirname(sync_path), "wineDB.db")
                ui_manager.add_text_content("Copying database file to sync location...")
                shutil.copy(db_output_path, db_sync_path)
                ui_manager.add_text_content(
                    f"\033[32mSpreadsheet synced to drive location: {sync_path}\033[39m")
                ui_manager.add_text_content(
                    f"\033[32mDatabase file synced to drive location: {db_sync_path}\033[39m")
        except Exception as exc:
            logger.error("Error loading settings from %s: %s", SETTINGS_FILE_PATH, str(exc))
            ui_manager.add_text_content(
                "\033[31mError loading settings. See logs for details.\033[39m")

    time.sleep(3)
    return True
